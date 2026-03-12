from pathlib import Path
from collections import defaultdict
from datetime import datetime
import csv
import json


# --------------------------------------------------
# OPTIONAL XLSX SUPPORT
# --------------------------------------------------

try:
    from openpyxl import load_workbook
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


# --------------------------------------------------
# UNIVERSAL COLUMN NAME OPTIONS
# --------------------------------------------------

COLUMN_ALIASES = {
    "order_id": [
        "order_id",
        "transaction_id",
        "invoice_id",
        "sale_id",
        "id",
    ],
    "order_date": [
        "order_date",
        "transaction_date",
        "date",
        "sale_date",
        "invoice_date",
    ],
    "customer_name": [
        "customer_name",
        "customer",
        "customer_id",
        "client_name",
        "client",
        "buyer",
    ],
    "product": [
        "product",
        "item",
        "product_name",
        "item_name",
        "service",
    ],
    "category": [
        "category",
        "product_category",
        "item_category",
        "department",
        "type",
    ],
    "quantity": [
        "quantity",
        "qty",
        "units",
        "count",
    ],
    "unit_price": [
        "unit_price",
        "price_per_unit",
        "price",
        "unit cost",
        "rate",
    ],
    "total_spent": [
        "total_spent",
        "total",
        "total_amount",
        "amount",
        "revenue",
        "sales",
        "line_total",
    ],
    "city": [
        "city",
        "location",
        "branch",
        "area",
        "region",
        "store_location",
    ],
}


SUPPORTED_EXTENSIONS = {".csv", ".tsv", ".txt", ".json", ".xlsx", ".xls"}


# --------------------------------------------------
# HELPERS
# --------------------------------------------------

def ensure_folder(folder_path: Path) -> None:
    """Create folder if it does not exist."""
    folder_path.mkdir(parents=True, exist_ok=True)


def clean_text(value) -> str:
    """Trim whitespace safely."""
    if value is None:
        return ""
    return str(value).strip()


def normalize_header(header: str) -> str:
    """Normalize CSV header for alias matching."""
    return clean_text(header).lower().replace("-", "_").replace(" ", "_")


def parse_float(value) -> float:
    """Convert text to float safely."""
    text = clean_text(value).replace(",", "")
    if text == "":
        raise ValueError("Empty numeric value")
    return float(text)


def parse_date(value) -> datetime:
    """Parse date using several common formats."""
    text = clean_text(value)

    date_formats = [
        "%Y-%m-%d",
        "%d-%m-%Y",
        "%m-%d-%Y",
        "%Y/%m/%d",
        "%d/%m/%Y",
        "%m/%d/%Y",
        "%Y-%m-%d %H:%M:%S",
        "%d-%m-%Y %H:%M:%S",
        "%m-%d-%Y %H:%M:%S",
        "%Y/%m/%d %H:%M:%S",
        "%d/%m/%Y %H:%M:%S",
        "%m/%d/%Y %H:%M:%S",
    ]

    for fmt in date_formats:
        try:
            return datetime.strptime(text, fmt)
        except ValueError:
            continue

    raise ValueError(f"Unsupported date format: {text}")


def build_header_map(fieldnames: list[str]) -> dict:
    """Build a map from standard field names to actual file headers."""
    normalized_to_original = {
        normalize_header(name): name for name in fieldnames
    }

    header_map = {}

    for standard_name, alias_list in COLUMN_ALIASES.items():
        for alias in alias_list:
            normalized_alias = normalize_header(alias)
            if normalized_alias in normalized_to_original:
                header_map[standard_name] = normalized_to_original[normalized_alias]
                break

    return header_map


def get_row_value(row: dict, header_map: dict, standard_key: str) -> str:
    """Get value from row using detected column mapping."""
    actual_column = header_map.get(standard_key)
    if not actual_column:
        return ""
    return clean_text(row.get(actual_column, ""))


# --------------------------------------------------
# FILE DISCOVERY
# --------------------------------------------------

def find_first_supported_file(input_folder: Path) -> Path:
    """Find the first supported input file inside input folder."""
    if not input_folder.exists():
        raise FileNotFoundError(f"Input folder not found: {input_folder}")

    all_files = sorted([p for p in input_folder.iterdir() if p.is_file()])

    if not all_files:
        raise FileNotFoundError(
            f"No input files found inside: {input_folder}"
        )

    for file_path in all_files:
        if file_path.suffix.lower() in SUPPORTED_EXTENSIONS:
            return file_path

    found_extensions = ", ".join(sorted({p.suffix.lower() or "[no extension]" for p in all_files}))
    raise ValueError(
        "No supported input file found.\n"
        f"Supported types: {', '.join(sorted(SUPPORTED_EXTENSIONS))}\n"
        f"Found types: {found_extensions}"
    )


# --------------------------------------------------
# FILE READERS
# --------------------------------------------------

def read_delimited_file(file_path: Path) -> list[dict]:
    """Read CSV/TSV/TXT files using delimiter detection."""
    with file_path.open("r", encoding="utf-8-sig", newline="") as f:
        sample = f.read(4096)
        f.seek(0)

        delimiter = ","
        if file_path.suffix.lower() == ".tsv":
            delimiter = "\t"
        else:
            try:
                sniffed = csv.Sniffer().sniff(sample, delimiters=",;\t|")
                delimiter = sniffed.delimiter
            except csv.Error:
                delimiter = ","

        reader = csv.DictReader(f, delimiter=delimiter)

        if not reader.fieldnames:
            raise ValueError("Delimited file has no headers.")

        return list(reader)


def read_json_file(file_path: Path) -> list[dict]:
    """Read JSON file as list of records."""
    with file_path.open("r", encoding="utf-8-sig") as f:
        data = json.load(f)

    if isinstance(data, list):
        if all(isinstance(item, dict) for item in data):
            return data
        raise ValueError("JSON list must contain objects/records.")

    if isinstance(data, dict):
        # Common case: {"data": [...]} or {"records": [...]}
        for key in ["data", "records", "rows", "items"]:
            value = data.get(key)
            if isinstance(value, list) and all(isinstance(item, dict) for item in value):
                return value

    raise ValueError("Unsupported JSON structure. Expected list of records.")


def read_xlsx_file(file_path: Path) -> list[dict]:
    """Read first sheet from XLSX file."""
    if not OPENPYXL_AVAILABLE:
        raise ValueError(
            "XLSX detected, but openpyxl is not installed. "
            "Install it with: pip install openpyxl"
        )

    workbook = load_workbook(file_path, data_only=True)
    sheet = workbook.active

    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        raise ValueError("XLSX file is empty.")

    headers = [clean_text(cell) for cell in rows[0]]
    if not any(headers):
        raise ValueError("XLSX file has no valid headers.")

    records = []
    for row_values in rows[1:]:
        row_dict = {}
        for index, header in enumerate(headers):
            if not header:
                continue
            value = row_values[index] if index < len(row_values) else ""
            row_dict[header] = value
        records.append(row_dict)

    return records


def load_raw_records(file_path: Path) -> list[dict]:
    """Read supported file types and return list of dict rows."""
    suffix = file_path.suffix.lower()

    if suffix in {".csv", ".tsv", ".txt"}:
        return read_delimited_file(file_path)

    if suffix == ".json":
        return read_json_file(file_path)

    if suffix == ".xlsx":
        return read_xlsx_file(file_path)

    if suffix == ".xls":
        raise ValueError(
            "Old .xls files are not supported in this version. "
            "Please save the file as .xlsx or .csv first."
        )

    raise ValueError(
        f"Unsupported file type: {suffix}. "
        f"Supported types: {', '.join(sorted(SUPPORTED_EXTENSIONS))}"
    )


# --------------------------------------------------
# LOADING + CLEANING
# --------------------------------------------------

def load_and_clean_data(input_file: Path) -> tuple[list[dict], int, dict]:
    """
    Load records from file, auto-detect columns, clean rows,
    and return:
    - cleaned rows
    - original row count
    - detected header map
    """
    raw_records = load_raw_records(input_file)

    if not raw_records:
        raise ValueError("Input file has no data rows.")

    first_row = raw_records[0]
    if not isinstance(first_row, dict):
        raise ValueError("Input data must contain row records as dictionaries.")

    header_map = build_header_map(list(first_row.keys()))

    required_minimum = ["order_id", "order_date", "product", "category"]
    missing_required = [key for key in required_minimum if key not in header_map]

    if missing_required:
        raise ValueError(
            "Could not detect required columns: "
            + ", ".join(missing_required)
        )

    cleaned_rows = []
    seen_rows = set()
    rows_before = 0

    for raw_row in raw_records:
        rows_before += 1

        try:
            order_id = get_row_value(raw_row, header_map, "order_id")
            order_date_raw = get_row_value(raw_row, header_map, "order_date")
            customer_name = get_row_value(raw_row, header_map, "customer_name")
            product = get_row_value(raw_row, header_map, "product")
            category = get_row_value(raw_row, header_map, "category")
            quantity_raw = get_row_value(raw_row, header_map, "quantity")
            unit_price_raw = get_row_value(raw_row, header_map, "unit_price")
            total_spent_raw = get_row_value(raw_row, header_map, "total_spent")
            city = get_row_value(raw_row, header_map, "city")

            order_date = parse_date(order_date_raw)

            quantity = None
            if quantity_raw != "":
                quantity = parse_float(quantity_raw)

            unit_price = None
            if unit_price_raw != "":
                unit_price = parse_float(unit_price_raw)

            revenue = None
            if total_spent_raw != "":
                revenue = parse_float(total_spent_raw)

            if revenue is None and quantity is not None and unit_price is not None:
                revenue = quantity * unit_price

            if quantity is None and revenue is not None and unit_price not in (None, 0):
                quantity = revenue / unit_price

            if unit_price is None and revenue is not None and quantity not in (None, 0):
                unit_price = revenue / quantity

            if quantity is None:
                quantity = 1.0

            if unit_price is None:
                unit_price = 0.0

            if revenue is None:
                revenue = quantity * unit_price

            row = {
                "order_id": order_id,
                "order_date": order_date,
                "customer_name": customer_name if customer_name else "Unknown",
                "product": product,
                "category": category,
                "quantity": quantity,
                "unit_price": unit_price,
                "city": city if city else "Unknown",
                "revenue": revenue,
                "month": order_date.strftime("%Y-%m"),
            }

        except ValueError:
            continue

        if (
            not row["order_id"]
            or not row["product"]
            or not row["category"]
            or row["quantity"] < 0
            or row["unit_price"] < 0
            or row["revenue"] < 0
        ):
            continue

        duplicate_key = (
            row["order_id"],
            row["order_date"].strftime("%Y-%m-%d"),
            row["customer_name"],
            row["product"],
            row["category"],
            round(row["quantity"], 4),
            round(row["unit_price"], 4),
            row["city"],
            round(row["revenue"], 4),
        )

        if duplicate_key in seen_rows:
            continue

        seen_rows.add(duplicate_key)
        cleaned_rows.append(row)

    return cleaned_rows, rows_before, header_map


# --------------------------------------------------
# SUMMARY
# --------------------------------------------------

def generate_summary(rows: list[dict]) -> dict:
    """Generate business summary metrics."""
    total_orders = len({row["order_id"] for row in rows})
    total_units_sold = sum(row["quantity"] for row in rows)
    total_revenue = sum(row["revenue"] for row in rows)

    revenue_by_order = defaultdict(float)
    revenue_by_product = defaultdict(float)
    revenue_by_category = defaultdict(float)
    revenue_by_city = defaultdict(float)
    revenue_by_month = defaultdict(float)

    for row in rows:
        revenue_by_order[row["order_id"]] += row["revenue"]
        revenue_by_product[row["product"]] += row["revenue"]
        revenue_by_category[row["category"]] += row["revenue"]
        revenue_by_city[row["city"]] += row["revenue"]
        revenue_by_month[row["month"]] += row["revenue"]

    average_order_value = (
        sum(revenue_by_order.values()) / len(revenue_by_order)
        if revenue_by_order
        else 0.0
    )

    top_products = sorted(
        revenue_by_product.items(),
        key=lambda item: item[1],
        reverse=True
    )[:5]

    top_categories = sorted(
        revenue_by_category.items(),
        key=lambda item: item[1],
        reverse=True
    )

    sales_by_city = sorted(
        revenue_by_city.items(),
        key=lambda item: item[1],
        reverse=True
    )

    monthly_sales = sorted(
        revenue_by_month.items(),
        key=lambda item: item[0]
    )

    return {
        "total_orders": total_orders,
        "total_units_sold": total_units_sold,
        "total_revenue": total_revenue,
        "average_order_value": average_order_value,
        "top_products": top_products,
        "top_categories": top_categories,
        "sales_by_city": sales_by_city,
        "monthly_sales": monthly_sales,
    }


# --------------------------------------------------
# EXPORTS
# --------------------------------------------------

def write_cleaned_csv(rows: list[dict], output_file: Path) -> None:
    """Export cleaned rows to CSV."""
    fieldnames = [
        "order_id",
        "order_date",
        "customer_name",
        "product",
        "category",
        "quantity",
        "unit_price",
        "city",
        "revenue",
        "month",
    ]

    with output_file.open("w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()

        for row in rows:
            export_row = {
                "order_id": row["order_id"],
                "order_date": row["order_date"].strftime("%Y-%m-%d"),
                "customer_name": row["customer_name"],
                "product": row["product"],
                "category": row["category"],
                "quantity": f"{row['quantity']:.2f}",
                "unit_price": f"{row['unit_price']:.2f}",
                "city": row["city"],
                "revenue": f"{row['revenue']:.2f}",
                "month": row["month"],
            }
            writer.writerow(export_row)


def write_text_report(summary: dict, output_file: Path, input_file: Path, header_map: dict) -> None:
    """Write summary report as text."""
    lines = []
    lines.append("SALES SUMMARY REPORT")
    lines.append("=" * 60)
    lines.append(f"Source File         : {input_file.name}")
    lines.append("")
    lines.append("DETECTED COLUMN MAPPING")
    lines.append("-" * 60)
    for standard_name, actual_name in header_map.items():
        lines.append(f"{standard_name} -> {actual_name}")
    lines.append("")
    lines.append(f"Total Orders        : {summary['total_orders']}")
    lines.append(f"Total Units Sold    : {summary['total_units_sold']:.2f}")
    lines.append(f"Total Revenue       : {summary['total_revenue']:.2f}")
    lines.append(f"Average Order Value : {summary['average_order_value']:.2f}")
    lines.append("")

    lines.append("TOP 5 PRODUCTS BY REVENUE")
    lines.append("-" * 60)
    for product, revenue in summary["top_products"]:
        lines.append(f"{product}: {revenue:.2f}")
    lines.append("")

    lines.append("CATEGORY SALES")
    lines.append("-" * 60)
    for category, revenue in summary["top_categories"]:
        lines.append(f"{category}: {revenue:.2f}")
    lines.append("")

    lines.append("CITY / LOCATION SALES")
    lines.append("-" * 60)
    for city, revenue in summary["sales_by_city"]:
        lines.append(f"{city}: {revenue:.2f}")
    lines.append("")

    lines.append("MONTHLY SALES")
    lines.append("-" * 60)
    for month, revenue in summary["monthly_sales"]:
        lines.append(f"{month}: {revenue:.2f}")

    output_file.write_text("\n".join(lines), encoding="utf-8")


def write_markdown_report(summary: dict, output_file: Path, input_file: Path) -> None:
    """Write markdown report for GitHub-friendly preview."""
    lines = []
    lines.append("# Sales Summary Report")
    lines.append("")
    lines.append(f"**Source File:** `{input_file.name}`")
    lines.append("")
    lines.append("## Overview")
    lines.append("")
    lines.append(f"- **Total Orders:** {summary['total_orders']}")
    lines.append(f"- **Total Units Sold:** {summary['total_units_sold']:.2f}")
    lines.append(f"- **Total Revenue:** {summary['total_revenue']:.2f}")
    lines.append(f"- **Average Order Value:** {summary['average_order_value']:.2f}")
    lines.append("")
    lines.append("## Top 5 Products by Revenue")
    lines.append("")

    for product, revenue in summary["top_products"]:
        lines.append(f"- **{product}**: {revenue:.2f}")

    lines.append("")
    lines.append("## Category Sales")
    lines.append("")

    for category, revenue in summary["top_categories"]:
        lines.append(f"- **{category}**: {revenue:.2f}")

    lines.append("")
    lines.append("## City / Location Sales")
    lines.append("")

    for city, revenue in summary["sales_by_city"]:
        lines.append(f"- **{city}**: {revenue:.2f}")

    lines.append("")
    lines.append("## Monthly Sales")
    lines.append("")

    for month, revenue in summary["monthly_sales"]:
        lines.append(f"- **{month}**: {revenue:.2f}")

    output_file.write_text("\n".join(lines), encoding="utf-8")


# --------------------------------------------------
# MAIN
# --------------------------------------------------

def main() -> None:
    project_root = Path(__file__).parent
    input_folder = project_root / "input"
    output_folder = project_root / "output"

    ensure_folder(input_folder)
    ensure_folder(output_folder)

    cleaned_output_file = output_folder / "cleaned_sales_data.csv"
    text_report_file = output_folder / "sales_summary_report.txt"
    markdown_report_file = output_folder / "sales_summary_report.md"

    print("📊 Universal Business Report Automation")
    print("-" * 60)

    try:
        input_file = find_first_supported_file(input_folder)
        print(f"Detected input file: {input_file.name}")

        rows, rows_before, header_map = load_and_clean_data(input_file)

        if not rows:
            print("No valid rows were found after cleaning.")
            print("Check whether your file actually contains sales-style tabular data.")
            return

        rows_after = len(rows)
        summary = generate_summary(rows)

        write_cleaned_csv(rows, cleaned_output_file)
        write_text_report(summary, text_report_file, input_file, header_map)
        write_markdown_report(summary, markdown_report_file, input_file)

        print(f"Rows before cleaning: {rows_before}")
        print(f"Rows after cleaning : {rows_after}")
        print("")
        print("Detected column mapping:")
        for standard_name, actual_name in header_map.items():
            print(f"  {standard_name} -> {actual_name}")
        print("")
        print("Files created:")
        print(f"1. Cleaned CSV       : {cleaned_output_file}")
        print(f"2. Text report       : {text_report_file}")
        print(f"3. Markdown report   : {markdown_report_file}")
        print("")
        print("✅ Automation complete.")

    except FileNotFoundError as e:
        print("Input file error:")
        print(e)

    except ValueError as e:
        print("Data/file format error:")
        print(e)

    except Exception as e:
        print("Unexpected error:")
        print(str(e))


if __name__ == "__main__":
    main()